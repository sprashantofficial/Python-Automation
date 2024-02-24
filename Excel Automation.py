from openpyxl import load_workbook

file_path = 'C:/Temp/1.xlsx'
wb = load_workbook(file_path)
sheet = wb.active

file_path = 'C:/Temp/2.xlsx'
wb2 = load_workbook(file_path)
sheet2 = wb2.active

#new_data = [['10', 'John', 'Doe', 'Male', 'United States', '28', '21/08/1994', '1234'],
#           ['11', 'Jane', 'Doe', 'Female', 'United States', '28', '21/06/1994', '1534']]

#for data in new_data:
#    print(type(data))
#    #sheet.append(data)

row_count = sheet2.max_row
column_count = sheet2.max_column

list_of_list = []

#print(sheet.cell(row=2, column=8).value)

for i in range(2, row_count+1):
    list = []
    for j in range(1, column_count+1):
        val = sheet2.cell(row=i, column=j).value
        list.append(val)
    list_of_list.append(list)

for data in list_of_list:
    sheet.append(data)

wb.save(filename='C:/Temp/updated_test.xlsx')